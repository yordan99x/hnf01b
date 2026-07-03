"""
OPM Call/Price Range Optimizer
===============================
Reusable end-to-end script: given a spreadsheet with P/N, Total Calls, and
DN Price columns, it:
  1. Searches for Call Range breakpoints (A/B/C/D, given LA=2-3, C1=1, C0=0
     are fixed) so that, across the FULL dataset:
       Count of P/N:  D < C < B < A < LA < C1 < C0   (increasing)
       Total Calls:   D > C > B > A > LA > C1 > C0   (decreasing)
  2. Searches for DN Price breakpoints (4 limits -> 5 bands: to L1, to L2,
     to L3, to L4, high value), trying to keep L1 at or as close as
     possible (from above, if infeasible) to a target cap, so that:
       Count of P/N:  band1 > band2 > band3 > band4 > band5  (decreasing)
       Total Calls:   band1 > band2 > band3 > band4 > band5  (decreasing)
  3. Builds a new workbook with the raw data, the breakpoints as editable
     blue inputs, and two live-formula pivot tables (Count of P/N, Total
     Calls) cross-tabbed by Price Range (rows) x Call Range (columns).

Usage
-----
python opm_optimizer.py INPUT.xlsx OUTPUT.xlsx \
    --sheet Sheet2 --pn-col "P/N" --price-col "DN Price" --calls-col "Sum of Tot Call" \
    --price-cap 10

If --sheet/--pn-col/--price-col/--calls-col are omitted, the script assumes
the first sheet and looks for columns named (case-insensitively, partial
match ok) "P/N", "Price", and "Call".
"""
import argparse
import sys
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
CALL_CATS = ["D", "C", "B", "A", "LA", "C1", "C0"]


# --------------------------------------------------------------------------
# Data loading
# --------------------------------------------------------------------------
def load_data(path, sheet=None, pn_col=None, price_col=None, calls_col=None):
    xls = pd.ExcelFile(path)
    sheet = sheet or xls.sheet_names[0]
    raw = pd.read_excel(path, sheet_name=sheet)

    def find_col(explicit, keyword):
        if explicit:
            return explicit
        for c in raw.columns:
            if keyword.lower() in str(c).lower():
                return c
        raise ValueError(f"Could not auto-detect a column matching '{keyword}'. "
                          f"Available columns: {list(raw.columns)}")

    pn_col = find_col(pn_col, "P/N")
    price_col = find_col(price_col, "Price")
    calls_col = find_col(calls_col, "Call")

    df = raw[[pn_col, price_col, calls_col]].copy()
    df.columns = ["PN", "Price", "Calls"]
    df["Price"] = pd.to_numeric(df["Price"], errors="coerce").fillna(0)
    df["Calls"] = pd.to_numeric(df["Calls"], errors="coerce").fillna(0).astype(int)
    return df


# --------------------------------------------------------------------------
# Call Range breakpoint search (A/B/C/D boundaries, given LA/C1/C0 fixed)
# --------------------------------------------------------------------------
def find_call_breakpoints(df):
    LA_count = int(df["Calls"].isin([2, 3]).sum())
    LA_sum = int(df.loc[df["Calls"].isin([2, 3]), "Calls"].sum())

    sub = df[df["Calls"] >= 4]
    if sub.empty:
        raise ValueError("No items with Calls >= 4; cannot build A/B/C/D bands.")
    calls_arr = sub["Calls"].values
    vals = np.sort(sub["Calls"].unique()).tolist()

    def band_stats(lo, hi):
        mask = (calls_arr >= lo) & (calls_arr <= hi)
        return int(mask.sum()), int(calls_arr[mask].sum())

    results = []
    for ai, a_max in enumerate(vals):
        cntA, sumA = band_stats(4, a_max)
        if cntA >= LA_count or sumA <= LA_sum:
            continue
        for bi in range(ai + 1, len(vals)):
            b_max = vals[bi]
            cntB, sumB = band_stats(a_max + 1, b_max)
            if cntB >= cntA or sumB <= sumA:
                continue
            for ci in range(bi + 1, len(vals)):
                c_max = vals[ci]
                cntC, sumC = band_stats(b_max + 1, c_max)
                if cntC >= cntB or sumC <= sumB:
                    continue
                cntD, sumD = band_stats(c_max + 1, vals[-1])
                if cntD == 0 or cntD >= cntC or sumD <= sumC:
                    continue
                minratio = min(cntB / cntA, cntC / cntB, cntD / cntC, cntA / LA_count)
                results.append((minratio, a_max, b_max, c_max))

    if not results:
        raise ValueError("No valid Call Range breakpoints found for this dataset "
                          "(the A/B/C/D/LA/C1/C0 monotonic pattern isn't achievable).")
    results.sort(key=lambda x: x[0])
    _, a_max, b_max, c_max = results[0]
    return a_max, b_max, c_max


# --------------------------------------------------------------------------
# Price Range breakpoint search (4 limits -> 5 bands)
# --------------------------------------------------------------------------
def find_price_breakpoints(df, price_cap=10, max_cap_search=200):
    pdata = df[df["Price"] > 0].copy()
    if pdata.empty:
        raise ValueError("No items with Price > 0; cannot build price bands.")

    vals, inv = np.unique(pdata["Price"].values, return_inverse=True)
    cnt_per_val = np.bincount(inv, minlength=len(vals))
    sum_per_val = np.bincount(inv, weights=pdata["Calls"].values, minlength=len(vals))
    cum_cnt = np.cumsum(cnt_per_val)
    cum_sum = np.cumsum(sum_per_val)
    N = len(vals)
    total_cnt = int(cum_cnt[-1])
    total_sum = float(cum_sum[-1])

    def cum_at(i):
        if i < 0:
            return 0, 0.0
        return int(cum_cnt[i]), float(cum_sum[i])

    def band(prev_idx, idx):
        pc, ps = cum_at(prev_idx)
        c, s = cum_at(idx)
        return c - pc, s - ps

    def greedy_max(prev_idx, prev_cnt, prev_sum, min_start, max_idx):
        lo, hi, best = min_start, max_idx, None
        while lo <= hi:
            mid = (lo + hi) // 2
            bc, bs = band(prev_idx, mid)
            if bc < prev_cnt and bs < prev_sum and bc > 0:
                best = mid
                lo = mid + 1
            else:
                hi = mid - 1
        return best

    def dfs_from(idx1, tries_budget=150):
        cnt1, sum1 = cum_at(idx1)
        if cnt1 == 0:
            return None
        max2 = greedy_max(idx1, cnt1, sum1, idx1 + 1, N - 2)
        if max2 is None:
            return None
        idx2, t2 = max2, 0
        while idx2 > idx1 and t2 < tries_budget:
            cnt2, sum2 = band(idx1, idx2)
            if 0 < cnt2 < cnt1 and sum2 < sum1:
                max3 = greedy_max(idx2, cnt2, sum2, idx2 + 1, N - 2)
                if max3 is not None:
                    idx3, t3 = max3, 0
                    while idx3 > idx2 and t3 < tries_budget:
                        cnt3, sum3 = band(idx2, idx3)
                        if 0 < cnt3 < cnt2 and sum3 < sum2:
                            max4 = greedy_max(idx3, cnt3, sum3, idx3 + 1, N - 1)
                            if max4 is not None:
                                idx4, t4 = max4, 0
                                while idx4 > idx3 and t4 < tries_budget:
                                    cnt4, sum4 = band(idx3, idx4)
                                    if 0 < cnt4 < cnt3 and sum4 < sum3:
                                        cnt5 = total_cnt - cum_at(idx4)[0]
                                        sum5 = total_sum - cum_at(idx4)[1]
                                        if 0 < cnt5 < cnt4 and sum5 < sum4:
                                            return (idx1, idx2, idx3, idx4)
                                    idx4 -= 1
                                    t4 += 1
                        idx3 -= 1
                        t3 += 1
            idx2 -= 1
            t2 += 1
        return None

    target_idx1 = int(np.searchsorted(cum_cnt, max(1, total_cnt // 5), side="left"))
    cap_idx_candidates = np.where(vals <= price_cap)[0]
    start_idx = cap_idx_candidates.max() if len(cap_idx_candidates) > 0 else target_idx1

    sol = None
    for off in range(0, max_cap_search):
        idx1 = start_idx + off
        if idx1 >= N - 3:
            break
        sol = dfs_from(idx1)
        if sol:
            break

    if not sol:
        raise ValueError("No valid Price Range breakpoints found for this dataset.")
    i1, i2, i3, i4 = sol
    return float(vals[i1]), float(vals[i2]), float(vals[i3]), float(vals[i4])


# --------------------------------------------------------------------------
# Workbook builder
# --------------------------------------------------------------------------
def build_workbook(df, call_bps, price_bps, output_path):
    a_max, b_max, c_max = call_bps
    p1, p2, p3, p4 = price_bps
    n = len(df)
    last_row = 1 + n

    wb = Workbook()
    ws = wb.active
    ws.title = "OPM_Optimized"

    blue = Font(name=FONT, color="0000FF")
    black = Font(name=FONT, color="000000")
    bold = Font(name=FONT, bold=True)
    header_fill = PatternFill("solid", start_color="D9E1F2")
    title_font = Font(name=FONT, bold=True, size=12)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["P/N", "Tot Call", "DN Price", "Call Range", "Price Range"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = bold
        c.fill = header_fill

    for idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.cell(row=idx, column=1, value=row.PN).font = black
        ws.cell(row=idx, column=2, value=int(row.Calls)).font = black
        ws.cell(row=idx, column=2).number_format = "#,##0"
        ws.cell(row=idx, column=3, value=float(row.Price)).font = black
        ws.cell(row=idx, column=3).number_format = "#,##0.00"
        ws.cell(row=idx, column=4,
                value=f'=IF(B{idx}=0,"C0",IF(B{idx}=1,"C1",IF(B{idx}<=3,"LA",'
                      f'IF(B{idx}<=$H$2,"A",IF(B{idx}<=$H$3,"B",IF(B{idx}<=$H$4,"C","D"))))))').font = black
        ws.cell(row=idx, column=5,
                value=f'=IF(OR(C{idx}="",C{idx}<=0),"",IF(C{idx}<=$K$2,"to "&$K$2,'
                      f'IF(C{idx}<=$K$3,"to "&$K$3,IF(C{idx}<=$K$4,"to "&$K$4,'
                      f'IF(C{idx}<=$K$5,"to "&$K$5,"high value")))))').font = black

    for col, w in zip("ABCDE", (16, 10, 11, 12, 12)):
        ws.column_dimensions[col].width = w

    ws["G1"] = "Call Range Breakpoints (Tot Call)"
    ws["G1"].font = title_font
    ws["G2"] = "A max (Very slow, starts at 4)"; ws["H2"] = a_max
    ws["G3"] = "B max (Slow moving)"; ws["H3"] = b_max
    ws["G4"] = "C max (Medium moving)"; ws["H4"] = c_max
    ws["G5"] = "D = above C max (Fast moving)"
    ws["G6"] = "LA = calls 2-3, C1 = calls 1, C0 = calls 0 (fixed categories)"
    for r in (2, 3, 4):
        ws.cell(row=r, column=8).font = blue
    for r in (2, 3, 4, 5, 6):
        ws.cell(row=r, column=7).font = Font(name=FONT, italic=(r >= 5))
    ws.column_dimensions["G"].width = 34
    ws.column_dimensions["H"].width = 10

    ws["J1"] = "DN Price Range Breakpoints"
    ws["J1"].font = title_font
    ws["J2"] = "Limit of range 1"; ws["K2"] = p1
    ws["J3"] = "Limit of range 2"; ws["K3"] = p2
    ws["J4"] = "Limit of range 3"; ws["K4"] = p3
    ws["J5"] = "Limit of range 4"; ws["K5"] = p4
    ws["J6"] = "Above Limit 4 = high value"
    for r in (2, 3, 4, 5):
        ws.cell(row=r, column=11).font = blue
        ws.cell(row=r, column=11).number_format = "0.00"
    for r in (2, 3, 4, 5, 6):
        ws.cell(row=r, column=10).font = Font(name=FONT, italic=(r == 6))
    ws.column_dimensions["J"].width = 22
    ws.column_dimensions["K"].width = 10

    n_excluded = int((df["Price"] <= 0).sum())
    if n_excluded:
        ws["G7"] = (f"Note: {n_excluded} line item(s) have DN Price <= 0 and are excluded "
                    f"from the Price Range classification and both pivot tables below.")
        ws["G7"].font = Font(name=FONT, italic=True, size=9, color="808080")

    data_rng_D = f"$D$2:$D${last_row}"
    data_rng_E = f"$E$2:$E${last_row}"
    data_rng_B = f"$B$2:$B${last_row}"

    def build_pivot(start_row, title, value_formula_maker):
        ws.cell(row=start_row, column=7, value=title).font = title_font
        hdr = start_row + 1
        ws.cell(row=hdr, column=7, value="Price Range \\ Call Range").font = bold
        ws.cell(row=hdr, column=7).fill = header_fill
        for j, cat in enumerate(CALL_CATS, start=8):
            c = ws.cell(row=hdr, column=j, value=cat)
            c.font = bold; c.fill = header_fill; c.alignment = Alignment(horizontal="center")
        gt_col = 8 + len(CALL_CATS)
        c = ws.cell(row=hdr, column=gt_col, value="Grand Total")
        c.font = bold; c.fill = header_fill

        price_labels = [f'="to "&$K$2', f'="to "&$K$3', f'="to "&$K$4', f'="to "&$K$5', "high value"]
        first_data_row = hdr + 1
        for i, lbl in enumerate(price_labels):
            r = first_data_row + i
            ws.cell(row=r, column=7, value=lbl).font = black
            for j, cat in enumerate(CALL_CATS, start=8):
                col_letter = get_column_letter(j)
                formula = value_formula_maker(col_letter, f"$G{r}", hdr)
                ws.cell(row=r, column=j, value=formula).font = black
                ws.cell(row=r, column=j).number_format = "#,##0"
            first_cat_col = get_column_letter(8)
            last_cat_col = get_column_letter(7 + len(CALL_CATS))
            ws.cell(row=r, column=gt_col, value=f"=SUM({first_cat_col}{r}:{last_cat_col}{r})").font = bold
            ws.cell(row=r, column=gt_col).number_format = "#,##0"

        gt_row = first_data_row + len(price_labels)
        ws.cell(row=gt_row, column=7, value="Grand Total").font = bold
        last_data_row = first_data_row + len(price_labels) - 1
        for j in range(8, gt_col + 1):
            col_letter = get_column_letter(j)
            ws.cell(row=gt_row, column=j,
                    value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})").font = bold
            ws.cell(row=gt_row, column=j).number_format = "#,##0"
        for r in range(hdr, gt_row + 1):
            for j in range(7, gt_col + 1):
                ws.cell(row=r, column=j).border = border
        return gt_row

    def make_count_formula(col_letter, price_ref, hdr_row):
        return f"=COUNTIFS({data_rng_D},{col_letter}${hdr_row},{data_rng_E},{price_ref})"

    def make_sum_formula(col_letter, price_ref, hdr_row):
        return f"=SUMIFS({data_rng_B},{data_rng_D},{col_letter}${hdr_row},{data_rng_E},{price_ref})"

    end1 = build_pivot(9, "PIVOT 1 : Count of P/N", make_count_formula)
    start2 = end1 + 3
    build_pivot(start2, "PIVOT 2 : Total Calls", make_sum_formula)

    for col in ["H", "I", "J", "K", "L", "M", "N", "O"]:
        ws.column_dimensions[col].width = 12

    ws.freeze_panes = "A2"
    wb.save(output_path)


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="Optimize Call Range / Price Range breakpoints and build OPM pivots.")
    ap.add_argument("input", help="Input .xlsx with P/N, Total Calls, DN Price columns")
    ap.add_argument("output", help="Output .xlsx to create")
    ap.add_argument("--sheet", default=None, help="Sheet name (default: first sheet)")
    ap.add_argument("--pn-col", default=None, help="P/N column name (default: auto-detect)")
    ap.add_argument("--price-col", default=None, help="DN Price column name (default: auto-detect)")
    ap.add_argument("--calls-col", default=None, help="Total Calls column name (default: auto-detect)")
    ap.add_argument("--price-cap", type=float, default=10,
                     help="Preferred max value for the first price limit (default: 10)")
    args = ap.parse_args()

    df = load_data(args.input, args.sheet, args.pn_col, args.price_col, args.calls_col)
    print(f"Loaded {len(df)} rows.")

    call_bps = find_call_breakpoints(df)
    print(f"Call Range breakpoints -> A max={call_bps[0]}, B max={call_bps[1]}, C max={call_bps[2]}")

    price_bps = find_price_breakpoints(df, price_cap=args.price_cap)
    print(f"Price Range breakpoints -> {price_bps[0]:.2f}, {price_bps[1]:.2f}, "
          f"{price_bps[2]:.2f}, {price_bps[3]:.2f}")

    build_workbook(df, call_bps, price_bps, args.output)
    print(f"Workbook written to {args.output}")


if __name__ == "__main__":
    main()
